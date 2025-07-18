from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from time import sleep
from openpyxl import load_workbook

navegador = webdriver.Chrome()
lista_produtos = []

maiorPreco = 0
menorPreço = float('inf')
mediaPreços = 0
soma = 0

url = "https://www.amazon.com.br/s?k=capacete"
pesquisa = "notebook 16gb RAM e 1TB SSD"

navegador.get(url)
seila = navegador.find_element(By.XPATH, '/html/body/div/div/a/img')
seila.click()

barra_pesquisa = navegador.find_element(By.ID, 'twotabsearchtextbox')
barra_pesquisa.send_keys(pesquisa)
barra_pesquisa.send_keys(Keys.ENTER)
sleep(2)

marcas = navegador.find_element(By.ID, 'filter-p_123')
marca = marcas.find_element(By.ID, 'p_123/219979')

marca_html = marca.get_attribute('outerHTML')
marca_bs = BeautifulSoup(marca_html, 'html.parser')
marca_desejada = marca_bs.text.strip()

opcao = marca.find_element(By.CLASS_NAME, 'a-list-item')
opcao.click()

site = BeautifulSoup(navegador.page_source, 'html.parser')

produtos = site.find_all('div', attrs={
    'class': 'a-section a-spacing-base'
})

while len(lista_produtos) < 20:
    for produto in produtos[:20]:
        nome = produto.find('h2', attrs={
            'class': 'a-size-base-plus a-spacing-none a-color-base a-text-normal'
        })

        preco = produto.find('span', attrs={
            'class': 'a-price-whole'
        })

        avaliacao = produto.find('span', attrs={
            'class': 'a-icon-alt'
        })

        if nome and preco != None and avaliacao:
            precoValor = float(preco.text.strip().replace('R$', '').replace('.', '').replace(',', '.').strip())

            soma += precoValor

            if precoValor > maiorPreco:
                maiorPreco = precoValor
            if precoValor < menorPreço:
                menorPreço = precoValor

            item = {
                'marca': marca_desejada,
                'nome': nome.text,
                'preco': precoValor,
                'avaliacao': avaliacao.text
            }

            lista_produtos.append(item)
            print(item)

        if len(lista_produtos) >= 20:
            break

print("\n\nTamanho da lista: ", len(lista_produtos))

if lista_produtos: 
    mediaPreços = soma / len(lista_produtos)
    print("\n\nMaior preço: ", maiorPreco)
    print("Menor preço: ", menorPreço)
    print("Média de preços: ", mediaPreços)

arq_ex = "Produtos.xlsx"

wb = load_workbook(arq_ex)
ws_amazon = wb.create_sheet(title="Amazon Produtos")
ws_amazon.append(['Marca', 'Nome', 'Preço', 'Avaliação'])


for produto in lista_produtos:
    ws_amazon.append([produto['marca'], produto['nome'], produto['preco'], produto['avaliacao']])
    
precos_amazon = wb.create_sheet(title="Preços Amazon")
precos_amazon.append(['Maior Preço', 'Menor Preço', 'Média Preços'])
precos_amazon.append([maiorPreco, menorPreço, mediaPreços])

wb.save(arq_ex)
navegador.quit()